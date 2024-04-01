#!/usr/bin/env python3

import io
import logging
import os
import os.path

import openpyxl
import orjson
import requests
from pywebio import start_server
from pywebio.input import actions, file_upload, input_group
from pywebio.output import clear, put_button, put_error, put_success, put_text
from pywebio.session import run_js
from datetime import datetime, time, date

logging.basicConfig(level=os.getenv("LOGGING", "INFO"))

PROCEDURE_URL = os.getenv(
    "PROCEDURE_URL", "http://localhost:8000/hasura/api/rest/import-data"
)

logging.debug(os.environ)


elvedata_mapping = {
    "Start dato": "start_dato",
    "Slutt dato": "slutt_dato",
    "Elv": "elv",
    "Båttype": "baattype",
    "Lat": "lat",
    "Long": "lon",
    "Vannføring (sildre.no)": "vannfoering",
    "Skipper": "skipper",
    "Mannskap1": "mannskap1",
    "Mannskap2": "mannskap2",
    "Mannskap3": "mannskap3",
    "Prosjekt": "prosjektnavn",
    "Prosjektnummer": "prosjektnummer",
    "Kommentar": "kommentar",
}

stasjonsdata_mapping = {
    "Stasjon": "stasjonnummer",
    "Båttype": "baattype",
    "Dato": "dato",
    "Klokkeslett start": "klokkeslett_start",
    "Lat start": "lat_start",
    "Long start": "lon_start",
    "Lat stopp": "lat_stopp",
    "Long stopp": "lon_stopp",
    "Dominerende elvetype": "dominerende_elvetype",
    "Vær": "vaer",
    "Vanntemp (Celsius)": "vanntemp",
    "Lufttemperatur (Celsius)": "lufttemperatur",
    "Ledningsevne (µs/cm)": "ledningsevne",
    "Transektlengde (m)": "transektlengde",
    "Sekunder fisket (s)": "sekunder_fisket",
    "Volt": "volt",
    "Puls (DC)": "puls",
    "Display": "display",
    "Gpx file?": "gpx_file",
    "Stasjonsbeskrivelse": "stasjonsbeskrivelse",
    "Kommentarer til fiske (observasjoner osv)": "kommentar",
}

individdata_mapping = {
    "ID": "id",
    "Stasjon": "stasjon",
    "Omgang": "omgang",
    "Art": "art",
    "Lengde": "lengde",
    "Antall": "antall",
    "Kjønn": "kjoenn",
    "Alder": "alder",
    "Gjenutsatt (ja/nei)": "gjenutsatt",
    "Prøvetatt (ja/nei)": "proevetype",
    "Kommentar": "kommentar",
}

def create_daterange(start_date, end_date):
    # create a daterange string for Postgres
    start = start_date.strftime("%Y-%m-%d")
    end = end_date.strftime("%Y-%m-%d")
    return f"[{start}, {end}]"

def wizard():
    user_inputs = input_group(
        "Import",
        [
            file_upload("Select spreadsheets:", multiple=True, name="files"),
        ],
    )

    # Load Elvedata worksheet
    data = []
    for index, file in enumerate(user_inputs["files"]):
        workbook = openpyxl.load_workbook(io.BytesIO(file["content"]), data_only=True)
        # Elvedata
        rows = workbook["Elvedata"].iter_rows()
        header = [cell.value for cell in next(rows) if cell.value]
        logging.debug(header)
        header = [elvedata_mapping[column] for column in header]
        for row in rows:  # TODO: max 1 row!
            row = [cell.value for cell in row]
            logging.debug(row)
            if not any(row):
                continue
            elvedata = dict(zip(header, row))
            # start/enddate -> dato
            elvedata["dato"] = create_daterange(elvedata["start_dato"], elvedata["slutt_dato"])
            del elvedata["start_dato"]
            del elvedata["slutt_dato"] 
            # lat/lon -> posisjon
            elvedata["lon"] = float(elvedata["lon"])
            elvedata["lat"] = float(elvedata["lat"])
            elvedata["posisjon"] = {
                "type": "Point",
                "coordinates": [elvedata["lon"], elvedata["lat"]],
            }
            del elvedata["lon"]
            del elvedata["lat"]
            # mannskap* -> mannskap
            mannskap = []
            for index in range(1, 4):
                column_name = "mannskap%d" % index
                mannskap_value = elvedata[column_name]
                del elvedata[column_name]
                if not mannskap_value:
                    continue
                mannskap.append(mannskap_value)
            elvedata["mannskap"] = mannskap
            # Prepare and append
            elvedata["prosjektnummer"] = str(elvedata["prosjektnummer"])
            elvedata["stasjonsdata"] = {"data": []}
            data.append(elvedata)
        # Stasjonsdata
        rows = workbook["Stasjonsdata"].iter_rows()
        header = [cell.value for cell in next(rows) if cell.value]
        logging.debug(header)
        header = [stasjonsdata_mapping[column] for column in header]
        stasjoner = []
        for row in rows:
            row = [cell.value for cell in row]
            logging.debug(row)
            if not any(row):
                continue
            stasjonsdata = dict(zip(header, row))
            # time + date -> klokkeslett_start
            stasjonsdata["klokkeslett_start"] = datetime.combine(stasjonsdata["dato"], stasjonsdata["klokkeslett_start"])
            # lat/lon -> posisjon
            for suffix in ["_start", "_stopp"]:
                stasjonsdata["lon" + suffix] = float(stasjonsdata["lon" + suffix])
                stasjonsdata["lat" + suffix] = float(stasjonsdata["lat" + suffix])
                stasjonsdata["posisjon" + suffix] = {
                    "type": "Point",
                    "coordinates": [
                        stasjonsdata["lon" + suffix],
                        stasjonsdata["lat" + suffix],
                    ],
                }
                del stasjonsdata["lon" + suffix]
                del stasjonsdata["lat" + suffix]
            if stasjonsdata["display"] == "na":
                stasjonsdata["display"] = None
            # Prepare
            stasjoner.append(stasjonsdata["stasjonnummer"])
            stasjonsdata["individdata"] = {"data": []}
            # Drop useless columns
            del stasjonsdata["baattype"]
            del stasjonsdata["dato"]
            # Store row
            data[0]["stasjonsdata"]["data"].append(stasjonsdata)
        # Individdata
        rows = workbook["Individdata"].iter_rows()
        header = [cell.value for cell in next(rows) if cell.value]
        logging.debug(header)
        header = [individdata_mapping[column] for column in header]
        for row in rows:
            row = [cell.value for cell in row]
            logging.debug(row)
            if not any(row):
                continue
            individdata = dict(zip(header, row))
            # Boolean
            for column in ["gjenutsatt"]:
                if not individdata[column]:
                    continue
                individdata[column] = individdata[column].lower() == "ja"
            # Prepare
            stasjon_index = stasjoner.index(individdata["stasjon"])
            # Drop useless columns
            del individdata["stasjon"]
            del individdata["id"]
            # Store row
            data[0]["stasjonsdata"]["data"][stasjon_index]["individdata"][
                "data"
            ].append(individdata)
        logging.debug(data)

    put_text("The files have loaded.")
    actions(buttons=[{"label": "Import", "value": "import", "color": "primary"}])

    clear("import")
    put_text("Importing...")

    headers = {
        "X-Hasura-Admin-Secret": "admin",
    }
    try:
        headers["Content-Type"] = "application/json"
        data = orjson.dumps({"data": data})
        logging.debug(data)
        response = requests.post(
            PROCEDURE_URL,
            headers=headers,
            data=data,
        )
        logging.debug(response.text)
        response.raise_for_status()
    except Exception as instance:
        put_error(str(instance) + "\n" + response.text)
    else:
        put_success("Data has been imported sucessfully.")
    finally:
        put_button("Upload new data", onclick=lambda: run_js("window.location.reload()"))


if __name__ == "__main__":
    start_server(wizard, port=8000, debug=True)
